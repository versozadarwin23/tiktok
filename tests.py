import json

import os

import sys

import atexit

import string

import threading

from openpyxl import Workbook, load_workbook

import requests

from bs4 import BeautifulSoup

import time

import random

from zipfile import BadZipFile

# --- App Version and Update URL ---

__version__ = "V7"

UPDATE_URL = "https://raw.githubusercontent.com/versozadarwin23/tiktok/refs/heads/main/tests.py"

VERSION_CHECK_URL = "https://raw.githubusercontent.com/versozadarwin23/tiktok/refs/heads/main/version.txt"

ua = [
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921445772;FBDM/{density=4.0,width=1440,height=3120};FBLC/en_US;FBRV/0;FBCR/Verizon;FBMF/Samsung;FBBD/samsung;FBPN/com.facebook.katana;FBDV/SM-S948B;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921882103;FBDM/{density=3.0,width=1080,height=2400};FBLC/en_GB;FBRV/0;FBCR/Vodafone;FBMF/Xiaomi;FBBD/xiaomi;FBPN/com.facebook.katana;FBDV/25113PN0EG;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/580.1.0.44.110;FBBV/905331224;FBDM/{density=3.5,width=1440,height=3120};FBLC/id_ID;FBRV/0;FBCR/Telkomsel;FBMF/Google;FBBD/google;FBPN/com.facebook.katana;FBDV/Pixel 10 Pro;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921334551;FBDM/{density=4.0,width=1440,height=3120};FBLC/pt_BR;FBRV/0;FBCR/Vivo;FBMF/Oppo;FBBD/oppo;FBPN/com.facebook.katana;FBDV/CPH2841;FBSV/15;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/578.0.0.35.102;FBBV/892110447;FBDM/{density=3.0,width=1080,height=2400};FBLC/vi_VN;FBRV/0;FBCR/Viettel;FBMF/Vivo;FBBD/vivo;FBPN/com.facebook.katana;FBDV/V2502A;FBSV/15;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921990123;FBDM/{density=3.0,width=1080,height=2340};FBLC/es_ES;FBRV/0;FBCR/Movistar;FBMF/Samsung;FBBD/samsung;FBPN/com.facebook.katana;FBDV/SM-S942B;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/582.0.0.48.115;FBBV/912556778;FBDM/{density=3.5,width=1080,height=2400};FBLC/fr_FR;FBRV/0;FBCR/Orange;FBMF/Realme;FBBD/realme;FBPN/com.facebook.katana;FBDV/RMX3851;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921667112;FBDM/{density=4.0,width=1440,height=3200};FBLC/en_US;FBRV/0;FBCR/AT&T;FBMF/Samsung;FBBD/samsung;FBPN/com.facebook.katana;FBDV/SM-S948B;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/580.1.0.44.110;FBBV/905778223;FBDM/{density=3.0,width=1080,height=2400};FBLC/tr_TR;FBRV/0;FBCR/Turkcell;FBMF/Xiaomi;FBBD/xiaomi;FBPN/com.facebook.katana;FBDV/25113PN0EG;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921443119;FBDM/{density=3.5,width=1440,height=3120};FBLC/de_DE;FBRV/0;FBCR/Telekom;FBMF/Google;FBBD/google;FBPN/com.facebook.katana;FBDV/Pixel 10 Pro;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921102938;FBDM/{density=4.0,width=1440,height=3120};FBLC/en_PH;FBRV/0;FBCR/Globe;FBMF/Samsung;FBBD/samsung;FBPN/com.facebook.katana;FBDV/SM-S948B;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/582.0.0.48.115;FBBV/912883471;FBDM/{density=3.0,width=1080,height=2400};FBLC/en_AU;FBRV/0;FBCR/Telstra;FBMF/Xiaomi;FBBD/xiaomi;FBPN/com.facebook.katana;FBDV/25113PN0EG;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921556002;FBDM/{density=3.5,width=1440,height=3120};FBLC/en_CA;FBRV/0;FBCR/Rogers;FBMF/Google;FBBD/google;FBPN/com.facebook.katana;FBDV/Pixel 10 Pro;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/580.1.0.44.110;FBBV/905991234;FBDM/{density=4.0,width=1440,height=3120};FBLC/pt_PT;FBRV/0;FBCR/MEO;FBMF/Oppo;FBBD/oppo;FBPN/com.facebook.katana;FBDV/CPH2841;FBSV/15;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921778339;FBDM/{density=3.0,width=1080,height=2400};FBLC/it_IT;FBRV/0;FBCR/TIM;FBMF/Vivo;FBBD/vivo;FBPN/com.facebook.katana;FBDV/V2502A;FBSV/15;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/578.0.0.35.102;FBBV/892445110;FBDM/{density=3.0,width=1080,height=2340};FBLC/en_MY;FBRV/0;FBCR/Maxis;FBMF/Samsung;FBBD/samsung;FBPN/com.facebook.katana;FBDV/SM-S942B;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921221556;FBDM/{density=3.5,width=1080,height=2400};FBLC/en_ZA;FBRV/0;FBCR/MTN;FBMF/Realme;FBBD/realme;FBPN/com.facebook.katana;FBDV/RMX3851;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/582.0.0.48.115;FBBV/912334889;FBDM/{density=4.0,width=1440,height=3200};FBLC/nl_NL;FBRV/0;FBCR/KPN;FBMF/Samsung;FBBD/samsung;FBPN/com.facebook.katana;FBDV/SM-S948B;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921445221;FBDM/{density=3.0,width=1080,height=2400};FBLC/pl_PL;FBRV/0;FBCR/Play;FBMF/Xiaomi;FBBD/xiaomi;FBPN/com.facebook.katana;FBDV/25113PN0EG;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/580.1.0.44.110;FBBV/905112667;FBDM/{density=3.5,width=1440,height=3120};FBLC/en_IE;FBRV/0;FBCR/Three;FBMF/Google;FBBD/google;FBPN/com.facebook.katana;FBDV/Pixel 10 Pro;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921889443;FBDM/{density=4.0,width=1440,height=3120};FBLC/en_SG;FBRV/0;FBCR/Singtel;FBMF/Oppo;FBBD/oppo;FBPN/com.facebook.katana;FBDV/CPH2841;FBSV/15;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/578.0.0.35.102;FBBV/892776332;FBDM/{density=3.0,width=1080,height=2400};FBLC/ar_SA;FBRV/0;FBCR/STC;FBMF/Vivo;FBBD/vivo;FBPN/com.facebook.katana;FBDV/V2502A;FBSV/15;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921331009;FBDM/{density=3.0,width=1080,height=2340};FBLC/en_NZ;FBRV/0;FBCR/Spark;FBMF/Samsung;FBBD/samsung;FBPN/com.facebook.katana;FBDV/SM-S942B;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/582.0.0.48.115;FBBV/912554221;FBDM/{density=3.5,width=1080,height=2400};FBLC/ja_JP;FBRV/0;FBCR/Docomo;FBMF/Realme;FBBD/realme;FBPN/com.facebook.katana;FBDV/RMX3851;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921669887;FBDM/{density=4.0,width=1440,height=3200};FBLC/ko_KR;FBRV/0;FBCR/SKT;FBMF/Samsung;FBBD/samsung;FBPN/com.facebook.katana;FBDV/SM-S948B;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/580.1.0.44.110;FBBV/905330112;FBDM/{density=3.0,width=1080,height=2400};FBLC/en_HK;FBRV/0;FBCR/CSL;FBMF/Xiaomi;FBBD/xiaomi;FBPN/com.facebook.katana;FBDV/25113PN0EG;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921448554;FBDM/{density=3.5,width=1440,height=3120};FBLC/en_AE;FBRV/0;FBCR/Etisalat;FBMF/Google;FBBD/google;FBPN/com.facebook.katana;FBDV/Pixel 10 Pro;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/582.0.0.48.115;FBBV/912771003;FBDM/{density=4.0,width=1440,height=3120};FBLC/en_IL;FBRV/0;FBCR/Cellcom;FBMF/Oppo;FBBD/oppo;FBPN/com.facebook.katana;FBDV/CPH2841;FBSV/15;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921992445;FBDM/{density=3.0,width=1080,height=2400};FBLC/en_PK;FBRV/0;FBCR/Jazz;FBMF/Vivo;FBBD/vivo;FBPN/com.facebook.katana;FBDV/V2502A;FBSV/15;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/580.1.0.44.110;FBBV/905118776;FBDM/{density=3.0,width=1080,height=2340};FBLC/en_NG;FBRV/0;FBCR/MTN;FBMF/Samsung;FBBD/samsung;FBPN/com.facebook.katana;FBDV/SM-S942B;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921114002;FBDM/{density=3.5,width=1080,height=2400};FBLC/en_KE;FBRV/0;FBCR/Safaricom;FBMF/Realme;FBBD/realme;FBPN/com.facebook.katana;FBDV/RMX3851;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/578.0.0.35.102;FBBV/892885441;FBDM/{density=4.0,width=1440,height=3200};FBLC/en_GH;FBRV/0;FBCR/AirtelTigo;FBMF/Samsung;FBBD/samsung;FBPN/com.facebook.katana;FBDV/SM-S948B;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921332998;FBDM/{density=3.0,width=1080,height=2400};FBLC/en_BD;FBRV/0;FBCR/Grameenphone;FBMF/Xiaomi;FBBD/xiaomi;FBPN/com.facebook.katana;FBDV/25113PN0EG;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/582.0.0.48.115;FBBV/912441007;FBDM/{density=3.5,width=1440,height=3120};FBLC/en_LK;FBRV/0;FBCR/Dialog;FBMF/Google;FBBD/google;FBPN/com.facebook.katana;FBDV/Pixel 10 Pro;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921553112;FBDM/{density=4.0,width=1440,height=3120};FBLC/en_QA;FBRV/0;FBCR/Ooredoo;FBMF/Oppo;FBBD/oppo;FBPN/com.facebook.katana;FBDV/CPH2841;FBSV/15;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/580.1.0.44.110;FBBV/905770443;FBDM/{density=3.0,width=1080,height=2400};FBLC/en_KW;FBRV/0;FBCR/Zain;FBMF/Vivo;FBBD/vivo;FBPN/com.facebook.katana;FBDV/V2502A;FBSV/15;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921882110;FBDM/{density=3.0,width=1080,height=2340};FBLC/en_OM;FBRV/0;FBCR/Omantel;FBMF/Samsung;FBBD/samsung;FBPN/com.facebook.katana;FBDV/SM-S942B;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/582.0.0.48.115;FBBV/912115667;FBDM/{density=3.5,width=1080,height=2400};FBLC/en_JO;FBRV/0;FBCR/Umniah;FBMF/Realme;FBBD/realme;FBPN/com.facebook.katana;FBDV/RMX3851;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921338779;FBDM/{density=4.0,width=1440,height=3200};FBLC/en_EG;FBRV/0;FBCR/Vodafone;FBMF/Samsung;FBBD/samsung;FBPN/com.facebook.katana;FBDV/SM-S948B;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/580.1.0.44.110;FBBV/905441332;FBDM/{density=3.0,width=1080,height=2400};FBLC/en_MA;FBRV/0;FBCR/MarocTelecom;FBMF/Xiaomi;FBBD/xiaomi;FBPN/com.facebook.katana;FBDV/25113PN0EG;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921663001;FBDM/{density=3.5,width=1440,height=3120};FBLC/en_DZ;FBRV/0;FBCR/Mobilis;FBMF/Google;FBBD/google;FBPN/com.facebook.katana;FBDV/Pixel 10 Pro;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/582.0.0.48.115;FBBV/912998445;FBDM/{density=4.0,width=1440,height=3120};FBLC/en_TN;FBRV/0;FBCR/Ooredoo;FBMF/Oppo;FBBD/oppo;FBPN/com.facebook.katana;FBDV/CPH2841;FBSV/15;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921112338;FBDM/{density=3.0,width=1080,height=2400};FBLC/en_LB;FBRV/0;FBCR/Touch;FBMF/Vivo;FBBD/vivo;FBPN/com.facebook.katana;FBDV/V2502A;FBSV/15;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/580.1.0.44.110;FBBV/905335110;FBDM/{density=3.0,width=1080,height=2340};FBLC/en_ET;FBRV/0;FBCR/EthioTelecom;FBMF/Samsung;FBBD/samsung;FBPN/com.facebook.katana;FBDV/SM-S942B;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921557889;FBDM/{density=3.5,width=1080,height=2400};FBLC/en_UG;FBRV/0;FBCR/Airtel;FBMF/Realme;FBBD/realme;FBPN/com.facebook.katana;FBDV/RMX3851;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/582.0.0.48.115;FBBV/912224003;FBDM/{density=4.0,width=1440,height=3200};FBLC/en_TZ;FBRV/0;FBCR/Vodacom;FBMF/Samsung;FBBD/samsung;FBPN/com.facebook.katana;FBDV/SM-S948B;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921880112;FBDM/{density=3.0,width=1080,height=2400};FBLC/en_ZM;FBRV/0;FBCR/Zamtel;FBMF/Xiaomi;FBBD/xiaomi;FBPN/com.facebook.katana;FBDV/25113PN0EG;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/580.1.0.44.110;FBBV/905448997;FBDM/{density=3.5,width=1440,height=3120};FBLC/en_ZW;FBRV/0;FBCR/Econet;FBMF/Google;FBBD/google;FBPN/com.facebook.katana;FBDV/Pixel 10 Pro;FBSV/16;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/585.0.0.52.118;FBBV/921116445;FBDM/{density=4.0,width=1440,height=3120};FBLC/en_BW;FBRV/0;FBCR/Mascom;FBMF/Oppo;FBBD/oppo;FBPN/com.facebook.katana;FBDV/CPH2841;FBSV/15;FBOP/1;FBCA/arm64-v8a;]",
  "[FBAN/FB4A;FBAV/582.0.0.48.115;FBBV/912663110;FBDM/{density=3.0,width=1080,height=2400};FBLC/en_NA;FBRV/0;FBCR/MTC;FBMF/Vivo;FBBD/vivo;FBPN/com.facebook.katana;FBDV/V2502A;FBSV/15;FBOP/1;FBCA/arm64-v8a;]"
]

def clear_console():
    try:
        os.system("cls" if os.name == "nt" else "clear")
    except Exception:
        pass

def cprint(text, color=""):
    """Print text centered on the terminal, stripping ANSI for width calc."""
    import re
    ansi_escape = re.compile(r'\033\[[0-9;]*m')
    visible = ansi_escape.sub('', text)
    try:
        cols = os.get_terminal_size().columns
    except Exception:
        cols = 80
    pad = max(0, (cols - len(visible)) // 2)
    print(" " * pad + text)

def cinput(prompt):
    """Centered input prompt."""
    import re
    # Strip leading \n from prompt and print separately so centering is accurate
    clean_prompt = prompt.lstrip('\n')
    leading_newlines = len(prompt) - len(clean_prompt)
    if leading_newlines:
        print('\n' * (leading_newlines - 1))
    ansi_escape = re.compile(r'\033\[[0-9;]*m')
    visible = ansi_escape.sub('', clean_prompt)
    try:
        cols = os.get_terminal_size().columns
    except Exception:
        cols = 80
    pad = max(0, (cols - len(visible)) // 2)
    return input(" " * pad + clean_prompt)

def print_banner():
    clear_console()
    W = 42
    print()
    print()
    print()
    print()
    cprint("\033[1;34m" + "═" * W + "\033[0m")
    cprint("\033[1;33m⚡  FB Account Creator  ⚡\033[0m")
    cprint("\033[1;35mDeveloper  : Dars\033[0m")
    cprint("\033[1;32mVersion    : " + __version__ + "\033[0m")
    cprint("\033[1;34m" + "═" * W + "\033[0m")
    print()

def check_for_update():
    print_banner()
    cprint("\033[96m🔄 Checking for updates...\033[0m")

    try:
        response = requests.get(VERSION_CHECK_URL, timeout=5)
        response.raise_for_status()
        remote_version = response.text.strip()

        if remote_version != __version__:
            print()
            cprint("\033[1;93m╔══════════════════════════════════╗\033[0m")
            cprint("\033[1;93m║       UPDATE AVAILABLE!          ║\033[0m")
            cprint("\033[1;93m╚══════════════════════════════════╝\033[0m")
            cprint(f"\033[91mCurrent : {__version__}\033[0m")
            cprint(f"\033[92mLatest  : {remote_version}\033[0m")
            print()

            while True:
                choice = cinput("\033[93mUpdate now? (y/n): \033[0m").strip().lower()

                if choice == 'y':
                    cprint("\033[96m⬇️  Downloading update...\033[0m")
                    new_script_response = requests.get(UPDATE_URL, timeout=10)
                    new_script_response.raise_for_status()
                    current_script_name = sys.argv[0]
                    with open(current_script_name, 'w', encoding='utf-8') as f:
                        f.write(new_script_response.text)
                    cprint("\033[92m✅ Update successful! Please restart.\033[0m")
                    sys.exit()

                elif choice == 'n':
                    cprint("\033[91m❌ Update declined. Exiting...\033[0m")
                    sys.exit()

                else:
                    cprint("\033[91m⚠️  Type 'y' or 'n'.\033[0m")

        else:
            cprint(f"\033[92m✅ You're on the latest version ({__version__}).\033[0m")
            time.sleep(2)

    except requests.exceptions.RequestException as e:
        print()
        cprint(f"\033[91m❌ Update check failed: {e}\033[0m")
        cprint("\033[93m⚠️  Continuing with current version...\033[0m")
        time.sleep(2)

    except Exception as e:
        print()
        cprint(f"\033[91m❌ Unexpected error: {e}\033[0m")
        cprint("\033[93m⚠️  Continuing with current version...\033[0m")
        time.sleep(2)

check_for_update()

xlsx_lock    = threading.Lock()

console_lock = threading.Lock()

FAILURE  = "❌"

MAX_RETRIES = 3

CONFIG_FILE = "settings.json"

# ─────────────────────────────────────────────────────────────

# TEMPMAIL API  (your own server)

# ─────────────────────────────────────────────────────────────

TEMPMAIL_API = "https://signlang.oo.gd/tempmail/tm_public_api.php"

def generate_email():

    """

    Generate a disposable email using your own TempMail API.

    Returns (email_address, email_address) — second value kept for

    compatibility with the old (email, url) tuple callers.

    """

    for attempt in range(MAX_RETRIES):

        try:

            r = requests.get(TEMPMAIL_API, params={"action": "generate"}, timeout=15)

            r.raise_for_status()

            data = r.json()

            if data.get("success") and data.get("email"):

                email = data["email"]

                # Fix incomplete email domain (e.g. "user@signlang." missing "oo.gd")
                if email.endswith("@signlang.") or (email.count(".") == 0 and "@" in email):
                    email = email.rstrip(".") + ".oo.gd"
                elif "@" in email and not email.split("@")[1].count("."):
                    email = email + ".oo.gd"

                return email, email   # (email_address, inbox_ref)

        except Exception as e:

            with console_lock:

                cprint(f"\033[91m⚠️ generate_email attempt {attempt+1} failed: {e}\033[0m")

        time.sleep(2)

    raise RuntimeError("❌ Could not generate temp email after multiple attempts.")

def check_inbox_for_code(email, max_wait=120, poll_interval=5):

    """

    Poll your TempMail API inbox until a confirmation code arrives.

    Returns the code string, or None if timed out.

    """

    deadline = time.time() + max_wait

    local    = email.split("@")[0]

    while time.time() < deadline:

        try:

            r = requests.get(

                TEMPMAIL_API,

                params={"action": "inbox", "email": email},

                timeout=15

            )

            r.raise_for_status()

            data = r.json()

            if not data.get("success"):

                time.sleep(poll_interval)

                continue

            # latest_code is already extracted by the API

            if data.get("latest_code"):

                return data["latest_code"]

            # Fallback: scan subjects manually

            for msg in data.get("messages", []):

                subj = msg.get("subject", "")

                if "confirmation code" in subj.lower() or "is your" in subj.lower():

                    # Try to pull the code from the subject directly

                    import re

                    m = re.search(r'\b(\d{4,8})\b', subj)

                    if m:

                        return m.group(1)

                    # If not in subject, fetch the body

                    code = fetch_code_from_message(email, msg["id"])

                    if code:

                        return code

        except Exception as e:

            with console_lock:

                cprint(f"\033[91m⚠️ inbox poll error: {e}\033[0m")

        time.sleep(poll_interval)

    return None   # timed out

def fetch_code_from_message(email, msg_id):

    """Fetch a single message body and extract the OTP/code."""

    try:

        r = requests.get(

            TEMPMAIL_API,

            params={"action": "code", "email": email, "id": msg_id},

            timeout=15

        )

        r.raise_for_status()

        data = r.json()

        if data.get("success") and data.get("code"):

            return data["code"]

    except Exception:

        pass

    return None

# ─────────────────────────────────────────────────────────────

# DEVICE / UA HELPERS  (unchanged from original)

# ─────────────────────────────────────────────────────────────

def random_device_model():

    models = [

        "Samsung-SM-S918B","Xiaomi-2210132G","OnePlus-CPH2451","OPPO-CPH2207",

        "vivo-V2203","realme-RMX3085","Samsung-Galaxy-A54","Samsung-SM-A146P",

        "Samsung-Galaxy-S23Ultra","Samsung-SM-F946B","Samsung-Galaxy-M34",

        "Xiaomi-23049PCD8G","Xiaomi-Redmi-Note-12","Xiaomi-POCO-X5Pro",

        "Xiaomi-2312DRA50G","OnePlus-CPH2513","OnePlus-CPH2581","OnePlus-CPH2459",

        "OPPO-CPH2339","OPPO-CPH2419","OPPO-CPH2521","vivo-V2140","vivo-V2254",

        "vivo-V2230","vivo-V2313A","realme-RMX3612","realme-RMX3571",

        "realme-RMX3761","realme-RMX3491","Huawei-ANE-LX2","Huawei-JNY-LX1",

        "Huawei-ELS-NX9","Huawei-CDY-NX9B","Motorola-Moto-G73","Motorola-XT2345-4",

        "Motorola-XT2303-2","Infinix-X6815B","Infinix-X6711","Infinix-X676C",

        "TECNO-CK7n","TECNO-CH9n","TECNO-BD4h","HONOR-ANY-AN00","HONOR-MGA-AN00",

        "HONOR-LRA-AN00","Lenovo-L78051","Lenovo-K13-Note","Google-Pixel-7",

        "Google-Pixel-6a","Google-Pixel-5",

    ]

    return random.choice(models)

def random_device_id():

    ids = [

        "0f47e6d2-bb61-4bfc-80db-123456789001","1a2b3c4d-5e6f-7a8b-9c0d-234567890002",

        "2b3c4d5e-6f7a-8b9c-0d1e-345678900003","3c4d5e6f-7a8b-9c0d-1e2f-456789000004",

        "4d5e6f7a-8b9c-0d1e-2f3a-567890000005","5e6f7a8b-9c0d-1e2f-3a4b-678900000006",

        "6f7a8b9c-0d1e-2f3a-4b5c-789000000007","7a8b9c0d-1e2f-3a4b-5c6d-890000000008",

        "8b9c0d1e-2f3a-4b5c-6d7e-900000000009","9c0d1e2f-3a4b-5c6d-7e8f-000000000010",

    ]

    return random.choice(ids)

def random_fingerprint():

    fingerprints = [

        "samsung/a54/a54:13/TP1A.220624.014/A546EXXU1AWF2:user/release-keys",

        "samsung/s23ultra/s23ultra:14/UQ1A.240205.004/S918BXXU1AXBA:user/release-keys",

        "xiaomi/poco/poco:13/TKQ1.221013.002/V14.0.2.0.TKCMIXM:user/release-keys",

        "oneplus/CPH2513/CPH2513:14/UQ1A.240205.004/EX01:user/release-keys",

        "oppo/CPH2419/CPH2419:13/TP1A.220624.014/OP02:user/release-keys",

        "vivo/V2254/V2254:13/TP1A.220905.001/PD2254F_EX_A_13.1.5.7:user/release-keys",

        "realme/RMX3612/RMX3612:13/TP1A.220624.014/RMX3612_13_A.21:user/release-keys",

        "google/pixel7/pixel7:14/UQ1A.240205.004/10000001:user/release-keys",

        "motorola/XT2345-4/XT2345-4:13/TP1A.220624.014/20240403:user/release-keys",

        "huawei/CDY-NX9B/CDY-NX9B:11/HUAWEICDY-NX9B/678(user)/release-keys",

    ]

    return random.choice(fingerprints)

# ─────────────────────────────────────────────────────────────

# CONFIG / SETTINGS HELPERS

# ─────────────────────────────────────────────────────────────

def delete_config_file():

    if os.path.exists(CONFIG_FILE):

        try:

            os.remove(CONFIG_FILE)

        except Exception as e:

            print(f"⚠️ Failed to delete settings file: {e}")

atexit.register(delete_config_file)

def save_user_choice(key, value):

    data = {}

    if os.path.exists(CONFIG_FILE):

        with open(CONFIG_FILE, "r", encoding="utf-8") as f:

            try:

                data = json.load(f)

            except json.JSONDecodeError:

                data = {}

    data[key] = value

    with open(CONFIG_FILE, "w", encoding="utf-8") as f:

        json.dump(data, f, indent=2)

def load_user_choice(key):

    if not os.path.exists(CONFIG_FILE):

        return None

    with open(CONFIG_FILE, "r", encoding="utf-8") as f:

        try:

            return json.load(f).get(key)

        except json.JSONDecodeError:

            return None

# ─────────────────────────────────────────────────────────────

# FILE SAVE HELPERS

# ─────────────────────────────────────────────────────────────

def save_to_txt(filename, data):

    try:

        with open(filename, "a", encoding="utf-8") as f:

            f.write("|".join(data) + "\n")

    except Exception as e:

        print(f"\033[1;91m❗ Error saving to {filename}: {e}\033[0m")

def has_access_token_in_xlsx(filename, email_address):

    if not os.path.exists(filename):

        return False

    try:

        wb = load_workbook(filename)

    except BadZipFile:

        return False

    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):

        if len(row) > 4 and row[1] == email_address and row[4] and str(row[4]).strip():

            return True

    return False

def save_to_xlsx(filename, data):

    header_columns = ['NAME', 'USERNAME', 'PASSWORD', 'ACCOUNT LINK', 'ACCESS TOKEN']

    with xlsx_lock:

        while True:

            try:

                if os.path.exists(filename):

                    try:

                        wb = load_workbook(filename)

                        ws = wb.active

                    except BadZipFile:

                        os.remove(filename)

                        wb = Workbook()

                        ws = wb.active

                        ws.append(header_columns)

                else:

                    wb = Workbook()

                    ws = wb.active

                    ws.append(header_columns)

                if [cell.value for cell in ws[1]] != header_columns:

                    ws.delete_rows(1, ws.max_row)

                    ws.append(header_columns)

                existing_rows = [tuple(row) for row in ws.iter_rows(min_row=2, values_only=True)]

                if tuple(data) not in existing_rows:

                    ws.append(data)

                wb.save(filename)

                break

            except Exception as e:

                print(f"❗ Error saving to {filename}: {e}. Retrying in 1 second...")

                time.sleep(1)

# ─────────────────────────────────────────────────────────────

# NAME / USER DETAIL GENERATORS

# ─────────────────────────────────────────────────────────────

def load_names_from_file(file_path):

    try:

        with open(file_path, 'r', encoding='utf-8') as f:

            return [line.strip() for line in f if line.strip()]

    except FileNotFoundError:

        print(f"\033[91mError: {file_path} not found.\033[0m")

        return []

def get_names(account_type, gender):

    firstnames = load_names_from_file("first_name.txt")

    last_names  = load_names_from_file("last_name.txt")

    if not firstnames or not last_names:

        return "John", "Doe"

    return random.choice(firstnames), random.choice(last_names)

def generate_random_phone_number():

    return f"9{random.randint(0,4)}{random.randint(1,7)}{random.randint(1000000,9999999)}"

def generate_random_password():

    return 'Promises' + str(random.randint(100000, 999999))

def generate_user_details(account_type, gender, password=None):

    firstname, lastname = get_names(account_type, gender)

    year         = random.randint(1978, 2001)

    date         = random.randint(1, 28)

    month        = random.randint(1, 12)

    password     = password or generate_random_password()

    phone_number = generate_random_phone_number()

    return firstname, lastname, date, year, month, phone_number, password

# ─────────────────────────────────────────────────────────────

# GLOBALS

# ─────────────────────────────────────────────────────────────

custom_password_base = None

global_reg_choice    = load_user_choice("reg_choice")

# ─────────────────────────────────────────────────────────────

# MAIN ACCOUNT CREATOR

# ─────────────────────────────────────────────────────────────

def create_fbunconfirmed(account_type, usern, gender, password=None, session=None):

    global custom_password_base, global_reg_choice

    # ── Generate temp email from YOUR API ──────────────────

    email_address, _ = generate_email()

    with console_lock:

        cprint(f"\033[94m📧 Temp email: {email_address}\033[0m")

    agent = random.choice(ua)

    if password is None:

        if custom_password_base:

            password = custom_password_base + str(random.randint(100000, 999999))

        else:

            password = generate_random_password()

    firstname, lastname, date, year, month, phone_number, used_password = generate_user_details(account_type, gender, password)

    url = "https://m.facebook.com/reg"
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer": "https://m.facebook.com/reg",
        "Connection": "keep-alive",
        "Accept-Language": "en-US,en;q=0.9",
        "X-FB-Connection-Type": "mobile.LTE",
        "X-FB-Device": random_device_model(),
        "X-FB-Device-ID": random_device_id(),
        "X-FB-Fingerprint": random_fingerprint(),
        "X-FB-Connection-Quality": "EXCELLENT",
        "X-FB-Net-HNI": "51502",
        "X-FB-SIM-HNI": "51502",
        "X-FB-HTTP-Engine": "Liger",
        "x-fb-connection-type": "Unknown",
        "accept-encoding": "gzip, deflate",
        "content-type": "application/x-www-form-urlencoded",
        # "x-fb-http-engine": "Liger",
        "User-Agent": agent,
    }

    if session is None:

        session = requests.Session()

    # ── Get registration form ───────────────────────────────

    form = None

    for _ in range(MAX_RETRIES):

        try:

            def get_fb_form(url, headers):
                # Use a session to keep cookies (very important for 'wtsid' to work)
                session = requests.Session()

                # FIRST REQUEST
                resp = session.get(url, headers=headers, allow_redirects=False)

                # Loop to handle Facebook's jump-scares
                for _ in range(3):  # Try up to 3 redirects
                    if 'Location' in resp.headers:
                        new_url = resp.headers['Location']

                        # Fix the protocol crash
                        if new_url.startswith('fbredirect://'):
                            new_url = new_url.replace('fbredirect://', 'https://')

                        print(f"Following jump: {new_url}")

                        # Update Referer to satisfy security checks
                        headers['Referer'] = url
                        resp = session.get(new_url, headers=headers, allow_redirects=False)

                        # If we finally get a 200 OK, stop looping
                        if resp.status_code == 200:
                            break
                    else:
                        break

                return resp

            # Execute
            response = get_fb_form(url, headers)
            soup = BeautifulSoup(response.text, "html.parser")
            form = soup.find("form")

            if form:
                print("✅ Form successfully captured!")
                # Save the file for your debug
                with open("success_form.html", "w", encoding="utf-8") as f:
                    f.write(response.text)
            else:
                print("❌ Still no form. FB might be blocking based on IP/Header mismatch.")

            if form:

                break

            with console_lock:

                cprint("\033[91m😢 Registration form not found. Retrying...\033[0m")

        except requests.exceptions.RequestException as e:

            with console_lock:

                cprint(f"\033[91m😢 Network error getting form: {e}. Retrying...\033[0m")

        time.sleep(3)

    if not form:

        return "FAILED_NO_FORM"

    # ── Ask reg choice once ─────────────────────────────────

    if global_reg_choice is None:

        while True:
            print()
            cprint("\033[1;34m╔══════════════════════════════════╗\033[0m")
            cprint("\033[1;34m║     REGISTRATION METHOD          ║\033[0m")
            cprint("\033[1;34m╠══════════════════════════════════╣\033[0m")
            cprint("\033[1;34m║  \033[1;32m[1]\033[0m Email Registration          \033[1;34m║\033[0m")
            cprint("\033[1;34m║  \033[1;32m[2]\033[0m Phone Number Registration   \033[1;34m║\033[0m")
            cprint("\033[1;34m╚══════════════════════════════════╝\033[0m")

            choice = cinput("\n\033[1;92m➤ Your choice (1 or 2): \033[0m").strip()

            clear_console()

            if choice in ['1', '2']:
                global_reg_choice = choice
                save_user_choice("reg_choice", choice)
                break

            cprint("\033[91m❌ Invalid choice. Enter 1 or 2.\033[0m")

    else:

        choice = global_reg_choice

    is_phone_choice  = (choice == '2')

    email_or_phone   = phone_number if is_phone_choice else email_address

    if is_phone_choice:
        with console_lock:
            cprint(f"\033[92m📱 Using phone: {email_or_phone}\033[0m")

    # ── Submit registration ─────────────────────────────────

    data = {

        "firstname":      firstname,

        "lastname":       lastname,

        "birthday_day":   str(date),

        "birthday_month": str(month),

        "birthday_year":  str(year),

        "reg_email__":    email_or_phone,

        "sex":            str(gender),

        "encpass":        used_password,

        "submit":         "Sign Up",

    }

    action_url = requests.compat.urljoin(url, form.get("action", url))

    for inp in form.find_all("input"):

        if inp.has_attr("name") and inp["name"] not in data:

            data[inp["name"]] = inp.get("value", "")

    reg_response = None

    for _ in range(MAX_RETRIES):

        try:

            reg_response = session.post(action_url, headers=headers, data=data, timeout=60)

            reg_response.raise_for_status()

            break

        except requests.exceptions.RequestException as e:

            with console_lock:

                cprint(f"\033[91m😢 Registration post failed: {e}. Retrying...\033[0m")

            time.sleep(3)

    if not reg_response:

        return "FAILED_REGISTRATION"

    if "c_user" not in session.cookies:
        with console_lock:
            cprint("\033[91m⚠️ No c_user cookie. Account creation failed.\033[0m")
        time.sleep(3)
        return "FAILED_NO_C_USER"

    # ── Wait for confirmation code via YOUR TempMail API ────

    jbkj = check_inbox_for_code(email_address, max_wait=120, poll_interval=5)

    # ── If registered with phone, change to email ───────────

    if is_phone_choice:
        with console_lock:
            cprint("\033[93mChanging account email from phone...\033[0m")

        change_email_url = "https://m.facebook.com/changeemail/"

        change_form = None

        for _ in range(MAX_RETRIES):

            try:

                resp = session.get(change_email_url, headers=headers, timeout=60)

                soup = BeautifulSoup(resp.text, "html.parser")

                change_form = soup.find("form")

                if change_form:

                    break

            except requests.exceptions.RequestException as e:

                with console_lock:

                    cprint(f"\033[91m❌ Error getting email change form: {e}\033[0m")

            time.sleep(2)

        if change_form:

            action_url_change = requests.compat.urljoin(

                change_email_url, change_form.get("action", change_email_url)

            )

            data_change = {inp["name"]: inp.get("value", "")

                           for inp in change_form.find_all("input") if inp.has_attr("name")}

            data_change["new"]    = email_address

            data_change["submit"] = "Add"

            for _ in range(MAX_RETRIES):

                try:

                    cr = session.post(action_url_change, headers=headers, data=data_change, timeout=60)

                    cr.raise_for_status()

                    if "email" in cr.text.lower():
                        with console_lock:
                            cprint(f"\033[92m✅ Email changed to {email_address}\033[0m")
                        email_or_phone = email_address

                    break

                except requests.exceptions.RequestException as e:

                    with console_lock:

                        cprint(f"\033[91m❌ Email change post failed: {e}\033[0m")

                    time.sleep(3)

        # Re-check inbox after email change

        if not jbkj:
            with console_lock:
                cprint(f"\033[93m⏳ Re-checking inbox after email change...\033[0m")
            jbkj = check_inbox_for_code(email_address, max_wait=120, poll_interval=5)

    # ── Print result ────────────────────────────────────────

    full_name = f"{firstname} {lastname}"

    with console_lock:
        print()
        cprint("\033[1;34m╔════════════════════════════════════════╗\033[0m")
        cprint("\033[1;34m║        ✅ ACCOUNT CREATED              ║\033[0m")
        cprint("\033[1;34m╠════════════════════════════════════════╣\033[0m")
        cprint(f"\033[1;32m║  Name     : {full_name[:28]:<28}\033[1;34m║\033[0m")
        cprint(f"\033[1;32m║  Email    : {email_or_phone[:28]:<28}\033[1;34m║\033[0m")
        cprint(f"\033[1;32m║  Password : {used_password[:28]:<28}\033[1;34m║\033[0m")
        cprint(f"\033[1;32m║  Code     : {(jbkj if jbkj else 'N/A')[:28]:<28}\033[1;34m║\033[0m")
        cprint("\033[1;34m╚════════════════════════════════════════╝\033[0m")
        print()

    uid        = session.cookies.get("c_user")

    profile_id = f"https://www.facebook.com/profile.php?id={uid}"

    filename_xlsx = "/storage/emulated/0/Acc_Created.xlsx"

    filename_txt  = "/storage/emulated/0/Acc_created.txt"

    if not has_access_token_in_xlsx(filename_xlsx, email_or_phone):

        save_choice = cinput("\033[1;93m💾 Save this account? (y/n): \033[0m").strip().lower() or "y"

        if save_choice == "y":
            save_to_xlsx(filename_xlsx, [full_name, email_or_phone, used_password, profile_id, ""])
            save_to_txt(filename_txt,   [full_name, email_or_phone, used_password, profile_id, ""])
            with console_lock:
                cprint(f"\033[1;32m✅ Saved: {full_name}\033[0m")

# ─────────────────────────────────────────────────────────────

# ENTRY POINT

# ─────────────────────────────────────────────────────────────

def NEMAIN():

    print_banner()

    while True:
        try:
            cprint("\033[1;34m╔══════════════════════════════════╗\033[0m")
            cprint("\033[1;34m║        ACCOUNT CREATOR           ║\033[0m")
            cprint("\033[1;34m╚══════════════════════════════════╝\033[0m")

            print()
            count = int(cinput("\033[1;92m➤ How many accounts to create: \033[0m"))

            if count > 0:
                break

            cprint("\033[91m❌ Enter a positive number.\033[0m")

        except ValueError:
            cprint("\033[91m❌ Invalid input. Numbers only.\033[0m")

    global custom_password_base

    if custom_password_base is None:
        inp = cinput("\033[1;93m➤ Custom password base (blank = default): \033[0m").strip()
        custom_password_base = inp if inp else "Promises@"

    for _ in range(count):

        create_fbunconfirmed(1, "ali", 1, session=requests.Session())

if __name__ == "__main__":

    global_reg_choice = load_user_choice("reg_choice")

    if os.path.exists("settings.json"):

        os.remove("settings.json")

    global_reg_choice = None

    while True:

        NEMAIN()

