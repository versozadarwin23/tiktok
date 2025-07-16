import string
from openpyxl import Workbook, load_workbook
import os
import requests
from bs4 import BeautifulSoup
import time
import random
import json
import concurrent.futures
import threading
import hashlib
from flask import Flask, request, jsonify, render_template_string

app = Flask(__name__)

# --- Original functions from your script ---

def has_access_token_in_xlsx(filename, email_address):
    if not os.path.exists(filename):
        return False
    wb = load_workbook(filename)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        saved_email = row[1]
        saved_access_token = row[4]
        if saved_email == email_address and saved_access_token and saved_access_token.strip():
            return True
    return False

xlsx_lock = threading.Lock()
console_lock = threading.Lock()

def random_device_model():
    models = [
        "Samsung-SM-S918B", "Xiaomi-2210132G", "OnePlus-CPH2451", "OPPO-CPH2207",
        "vivo-V2203", "realme-RMX3085", "Samsung-Galaxy-A54", "Samsung-SM-A146P",
        "Samsung-Galaxy-S23Ultra", "Samsung-SM-F946B", "Samsung-Galaxy-M34",
        "Xiaomi-23049PCD8G", "Xiaomi-Redmi-Note-12", "Xiaomi-POCO-X5Pro",
        "Xiaomi-2312DRA50G", "OnePlus-CPH2513", "OnePlus-CPH2581",
        "OnePlus-CPH2459", "OPPO-CPH2339", "OPPO-CPH2419", "OPPO-CPH2521",
        "vivo-V2140", "vivo-V2254", "vivo-V2230", "vivo-V2313A",
        "realme-RMX3612", "realme-RMX3571", "realme-RMX3761",
        "realme-RMX3491", "Huawei-ANE-LX2", "Huawei-JNY-LX1",
        "Huawei-ELS-NX9", "Huawei-CDY-NX9B", "Motorola-Moto-G73",
        "Motorola-XT2345-4", "Motorola-XT2303-2", "Infinix-X6815B",
        "Infinix-X6711", "Infinix-X676C", "TECNO-CK7n", "TECNO-CH9n",
        "TECNO-BD4h", "HONOR-ANY-AN00", "HONOR-MGA-AN00", "HONOR-LRA-AN00",
        "Lenovo-L78051", "Lenovo-K13-Note", "Google-Pixel-7",
        "Google-Pixel-6a", "Google-Pixel-5"
    ]
    return random.choice(models)

def random_device_id():
    ids = [
        "0f47e6d2-bb61-4bfc-80db-123456789001", "1a2b3c4d-5e6f-7a8b-9c0d-234567890002",
        "2b3c4d5e-6f7a-8b9c-0d1e-345678900003", "3c4d5e6f-7a8b-9c0d-1e2f-456789000004",
        "4d5e6f7a-8b9c-0d1e-2f3a-567890000005", "5e6f7a8b-9c0d-1e2f-3a4b-678900000006",
        "6f7a8b9c-0d1e-2f3a-4b5c-789000000007", "7a8b9c0d-1e2f-3a4b-5c6d-890000000008",
        "8b9c0d1e-2f3a-4b5c-6d7e-900000000009", "9c0d1e2f-3a4b-5c6d-7e8f-000000000010",
        "aa1bb2cc-3dd4-5ee6-7ff8-111111111011", "bb2cc3dd-4ee5-6ff7-8009-222222222012",
        "cc3dd4ee-5ff6-7008-9110-333333333013", "dd4ee5ff-6007-8119-0221-444444444014",
        "ee5ff600-7118-9220-1332-555555555015", "ff600711-8229-0331-2443-666666666016",
        "00611722-9330-1442-3554-777777777017", "11722833-0441-2553-4665-888888888018",
        "22833944-1552-3664-5776-999999999019", "33944a55-2663-4775-6887-000000000020",
        "44a55b66-3774-5886-7998-111111111021", "55b66c77-4885-6997-8009-222222222022",
        "66c77d88-5996-7008-9110-333333333023", "77d88e99-6007-8119-0221-444444444024",
        "88e990aa-7118-9220-1332-555555555025", "990aa1bb-8229-0331-2443-666666666026",
        "0aa1bb2c-9330-1442-3554-777777777027", "1bb2cc3d-0441-2553-4665-888888888028",
        "2cc3dd4e-1552-3664-5776-999999999029", "3dd4ee5f-2663-4775-6887-000000000030",
        "4ee5ff60-3774-5886-7998-111111111031", "5ff60071-4885-6997-8009-222222222032",
        "60071182-5996-7008-9110-333333333033", "71182293-6007-8119-0221-444444444034",
        "82293304-7118-9220-1332-555555555045", "93304415-8229-0331-2443-666666666046",
        "04415526-9330-1442-3554-777777777047", "15526637-0441-2553-4665-888888888048",
        "26637748-1552-3664-5776-999999999049", "37748859-2663-4775-6887-000000000050"
    ]
    return random.choice(ids)

def random_fingerprint():
    fingerprints = [
        "samsung/a54/a54:13/TP1A.220624.014/A546EXXU1AWF2:user/release-keys",
        "samsung/m34/m34:13/TP1A.220624.014/M346BXXU1AWG3:user/release-keys",
        "samsung/s23ultra/s23ultra:14/UQ1A.240205.004/S918BXXU1AXBA:user/release-keys",
        "samsung/fold5/fold5:14/UQ1A.240205.004/F946BXXU1AWM7:user/release-keys",
        "xiaomi/umi/umi:12/RKQ1.211001.001/V12.5.6.0.RJBCNXM:user/release-keys",
        "xiaomi/poco/poco:13/TKQ1.221013.002/V14.0.2.0.TKCMIXM:user/release-keys",
        "xiaomi/redmi/redmi:14/UQ1A.240205.004/V14.0.5.0.ULOMIXM:user/release-keys",
        "xiaomi/note12/note12:13/TP1A.220624.014/V14.0.1.0.TKOMIXM:user/release-keys",
        "oneplus/CPH2513/CPH2513:14/UQ1A.240205.004/EX01:user/release-keys",
        "oneplus/CPH2451/CPH2451:13/TP1A.220905.001/EX02:user/release-keys",
        "oneplus/CPH2581/CPH2581:14/UQ1A.240205.004/EX03:user/release-keys",
        "oppo/CPH2207/CPH2207:12/SKQ1.211019.001/OP01:user/release-keys",
        "oppo/CPH2419/CPH2419:13/TP1A.220624.014/OP02:user/release-keys",
        "oppo/CPH2521/CPH2521:14/UQ1A.240205.004/OP03:user/release-keys",
        "vivo/V2203/V2203:12/SP1A.210812.016/PD2203F_EX_A_12.0.10.5:user/release-keys",
        "vivo/V2254/V2254:13/TP1A.220905.001/PD2254F_EX_A_13.1.5.7:user/release-keys",
        "vivo/V2313A/V2313A:14/UQ1A.240205.004/PD2313A_EX_A_14.0.3.2:user/release-keys",
        "realme/RMX3085/RMX3085:12/SP1A.210812.016/RMX3085_11_A.24:user/release-keys",
        "realme/RMX3612/RMX3612:13/TP1A.220624.014/RMX3612_13_A.21:user/release-keys",
        "realme/RMX3491/RMX3491:14/UQ1A.240205.004/RMX3491_14_A.11:user/release-keys",
        "huawei/ANE-LX2/ANE-LX2:10/HUAWEIANE-LX2/345(user)/release-keys",
        "huawei/CDY-NX9B/CDY-NX9B:11/HUAWEICDY-NX9B/678(user)/release-keys",
        "huawei/ELS-NX9/ELS-NX9:12/HUAWEIELS-NX9/901(user)/release-keys",
        "motorola/XT2345-4/XT2345-4:13/TP1A.220624.014/20240403:user/release-keys",
        "motorola/XT2303-2/XT2303-2:14/UQ1A.240205.004/20240501:user/release-keys",
        "infinix/X6815B/X6815B:12/SP1A.210812.016/X6815B-GL-220822V123:user/release-keys",
        "infinix/X6711/X6711:14/UQ1A.240205.004/X6711-GL-240104V101:user/release-keys",
        "infinix/X676C/X676C:13/TP1A.220624.014/X676C-H6120ABC-S-231015V104:user/release-keys",
        "tecno/CK7n/CK7n:14/UQ1A.240205.004/CK7n-H6121ABC-R-240305V103:user/release-keys",
        "tecno/CH9n/CH9n:13/TP1A.220624.014/CH9n-H6211ABC-R-231215V101:user/release-keys",
        "tecno/BD4h/BD4h:12/SP1A.210812.016/BD4h-H6112ABC-S-220915V102:user/release-keys",
        "honor/ANY-AN00/ANY-AN00:12/HONORANY-AN00/234(user)/release-keys",
        "honor/MGA-AN00/MGA-AN00:13/TP1A.220624.014/HONORMGA-AN00/567(user)/release-keys",
        "honor/LRA-AN00/LRA-AN00:14/UQ1A.240205.004/HONORLRA-AN00/890(user)/release-keys",
        "lenovo/L78051/L78051:12/SP1A.210812.016/L78051_USR_S_12.5.3:user/release-keys",
        "lenovo/K13-Note/K13-Note:13/TP1A.220624.014/K13Note_S_13.0.4:user/release-keys",
        "google/pixel7/pixel7:14/UQ1A.240205.004/10000001:user/release-keys",
        "google/pixel6a/pixel6a:14/UQ1A.240205.004/10000002:user/release-keys",
        "google/pixel5/pixel5:13/TP1A.220624.014/10000003:user/release-keys",
        "samsung/a146p/a146p:13/TP1A.220624.014/A146PXXU1AWF3:user/release-keys",
        "samsung/m54/m54:14/UQ1A.240205.004/M546BXXU1AXD2:user/release-keys",
        "xiaomi/2312DRA50G/2312DRA50G:14/UQ1A.240205.004/V14.0.7.0.UNOMIXM:user/release-keys",
        "xiaomi/23049PCD8G/23049PCD8G:13/TP1A.220624.014/V14.0.3.0.TMOMIXM:user/release-keys",
        "oneplus/CPH2459/CPH2459:14/UQ1A.240205.004/EX04:user/release-keys",
        "vivo/V2140/V2140:12/SP1A.210812.016/PD2140F_EX_A_12.0.9.8:user/release-keys",
        "realme/RMX3761/RMX3761:14/UQ1A.240205.004/RMX3761_14_A.13:user/release-keys",
        "motorola/Moto-G73/Moto-G73:13/TP1A.220624.014/20240401:user/release-keys",
        "infinix/X6711/X6711:14/UQ1A.240205.004/X6711-GL-240104V101:user/release-keys"
    ]
    return random.choice(fingerprints)

ua = [
    "Mozilla/5.0 (Linux; Android 10; SM-G960U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Pixel 5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; P30 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; SM-A525F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; Moto G Power) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Redmi Note 9S) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; OnePlus 9) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; LG G8 ThinQ) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; Xperia 5 II) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Pixel 4a) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Samsung SM-S901U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; ASUS_Z01QD) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; Vivo V2027) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Oppo A74) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Xiaomi 12) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; Nokia 7.2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; Realme 7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Infinix Note 10 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Tecno Camon 18) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; ZTE Axon 10 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; SM-A715F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Google Pixel 6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Samsung SM-G998B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; Huawei Mate 20 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; LG V60 ThinQ) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Samsung Galaxy A32) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Sony Xperia 1 III) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; Google Pixel 3a) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; OnePlus 8T) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Redmi K40 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Xiaomi 11 Lite 5G NE) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; Moto G7 Power) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; SM-G973U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Pixel 5a) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Samsung SM-A536B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; P40 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; Redmi Note 10 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; OnePlus Nord 2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; LG Wing) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; Xperia 1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; Asus ROG Phone 3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Vivo X70 Pro+) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Oppo Reno6 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; Nokia X20) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; Realme 8 Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Infinix Zero X Pro) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Tecno Pova 3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 9; ZTE Blade V2020) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/299.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 10; SM-A908B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/300.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 11; Google Pixel 7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/301.0.0.0.0;]",
    "Mozilla/5.0 (Linux; Android 12; Samsung SM-G990U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Mobile Safari/537.36 [FBAN/EMA;FBLC/en_US;FBAV/302.0.0.0.0;]"
]

def generate_email():
    rchjtrchjb = ''.join(random.choices(string.ascii_lowercase + string.digits, k=12))
    email_address = f"{rchjtrchjb}@harakirimail.com"
    drtyghbj5hgcbv = f"https://harakirimail.com/inbox/{rchjtrchjb}"
    return email_address, drtyghbj5hgcbv

def save_to_txt(filename, data):
    try:
        with open(filename, "a", encoding="utf-8") as f:
            f.write("|".join(data) + "\n")
    except Exception as e:
        print(f"❗ Error saving to {filename}: {e}")

def save_to_xlsx(filename, data):
    while True:
        with xlsx_lock:
            try:
                if os.path.exists(filename):
                    wb = load_workbook(filename)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.append(["NAME", "USERNAME", "PASSWORD", "ACCOUNT LINK", "ACCESS TOKEN"])
                ws.append(data)
                wb.save(filename)
                break
            except Exception as e:
                print(f"❗ Error saving to {filename}: {e}. Retrying...")
                time.sleep(RETRY_DELAY)

MAX_RETRIES = 3
RETRY_DELAY = 2
SUCCESS = "✅"
FAILURE = "❌"
INFO = "ℹ️"
WARNING = "⚠️"
LOADING = "⏳"

def load_names_from_file(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return [line.strip() for line in file.readlines() if line.strip()]
    except FileNotFoundError:
        return ["John", "Jane", "Doe", "Smith"]

def get_names(account_type, gender):
    male_first_names_file = "first_name.txt"
    last_names_file = "last_name.txt"

    male_first_names = load_names_from_file(male_first_names_file)
    last_names = load_names_from_file(last_names_file)

    if not male_first_names:
        male_first_names = ["Juan", "Pedro"]
    if not last_names:
        last_names = ["Dela Cruz", "Reyes"]

    firstname = random.choice(male_first_names)
    lastname = random.choice(last_names)
    return firstname, lastname

def generate_random_phone_number():
    random_number = str(random.randint(1000000, 9999999))
    third = random.randint(0, 4)
    forth = random.randint(1, 7)
    return f"9{third}{forth}{random_number}"

def generate_random_password():
    base = "Promises@"
    six_digit = str(random.randint(100000, 999999))
    return base + six_digit

def generate_user_details(account_type, gender, password=None):
    firstname, lastname, date, year, month, phone_number, used_password = (None,)*7 # Initialize to None to avoid UnboundLocalError
    firstname, lastname = get_names(account_type, gender)
    year = random.randint(1978, 2001)
    date = random.randint(1, 28)
    month = random.randint(1, 12)
    if password is None:
        password = generate_random_password()
    phone_number = generate_random_phone_number()
    return firstname, lastname, date, year, month, phone_number, password


custom_password_base = None

def create_fbunconfirmed_logic(account_num, account_type, gender, password=None, session=None):
    agent = random.choice(ua)
    global custom_password_base
    email_address, drtyghbj5hgcbv = generate_email()
    if password is None:
        if custom_password_base:
            six_digit = str(random.randint(100000, 999999))
            password = custom_password_base + six_digit
        else:
            password = generate_random_password()

    firstname, lastname, date, year, month, phone_number, used_password = generate_user_details(account_type, gender,
                                                                                                password)

    def check_page_loaded(url, headers, current_session):
        retries = 0
        while retries < MAX_RETRIES:
            try:
                response = current_session.get(url, timeout=30, headers=headers)
                soup = BeautifulSoup(response.text, 'html.parser')
                form = soup.find("form")
                if form:
                    return form
                else:
                    return None # Indicate form not found, but don't print here for Flask
            except requests.exceptions.RequestException as e:
                # Log errors but don't print directly to console for Flask
                pass
            except Exception as e:
                pass
            time.sleep(RETRY_DELAY)
            retries += 1
        return None

    url = "https://limited.facebook.com/reg?soft=hjk"
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer": "https://limited.facebook.com/reg?soft=hjk",
        "Connection": "keep-alive",
        "Accept-Language": "en-US,en;q=0.9",
        "X-FB-Connection-Type": "mobile.LTE",
        "X-FB-Device": random_device_model(),
        "X-FB-Device-ID": random_device_id(),
        "X-FB-Fingerprint": random_fingerprint(),
        "X-FB-Connection-Quality": "EXCELLENT",
        "X-FB-Net-HNI": "51502",
        "X-FB-SIM-HNI": "51502",
        "X-FB-HTTP-Engine": "Liger",
        'x-fb-connection-type': 'Unknown',
        'accept-encoding': 'gzip, deflate',
        'content-type': 'application/x-www-form-urlencoded',
        'User-Agent': agent,
    }
    if session is None:
        session = requests.Session()

    form = check_page_loaded(url, headers, session)
    if not form:
        return {"status": "FAILED_PAGE_LOAD", "message": f"Could not load registration page or find form. Aborting attempt for account #{account_num}."}

    retries = 0
    while retries < MAX_RETRIES:
        try:
            response = session.get(url, headers=headers, timeout=10)
            soup = BeautifulSoup(response.text, "html.parser")
            form = soup.find("form")
            if form:
                break
        except requests.exceptions.RequestException as e:
            pass
        except Exception as e:
            pass
        time.sleep(RETRY_DELAY)
        retries += 1

    if not form:
        return {"status": "FAILED_FORM_FETCH", "message": f"Failed to get registration form after retries. Aborting attempt for account #{account_num}."}

    action_url = requests.compat.urljoin(url, form["action"]) if form.has_attr("action") else url
    inputs = form.find_all("input")
    data = {
        "firstname": firstname,
        "lastname": lastname,
        "birthday_day": str(date),
        "birthday_month": str(month),
        "birthday_year": str(year),
        "reg_email__": email_address,
        "sex": str(gender),
        "encpass": used_password,
        "submit": "Sign Up"
    }

    for inp in inputs:
        if inp.has_attr("name") and inp["name"] not in data:
            data[inp["name"]] = inp.get("value", "")

    try:
        response = session.post(action_url, headers=headers, data=data, timeout=30)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        return {"status": "FAILED_SUBMISSION_NETWORK", "message": f"Network error during submission: {e}. Cannot complete account creation for account #{account_num}."}
    except Exception as e:
        return {"status": "FAILED_SUBMISSION_UNEXPECTED", "message": f"An unexpected error occurred during submission: {e}. Cannot complete account creation for account #{account_num}."}

    # For debugging, you might want to save this temporarily in a /tmp dir or similar
    # with open(f"status_{account_num}.html", "w", encoding="utf-8") as file:
    #    file.write(response.text)

    if "c_user" not in session.cookies:
        return {"status": "FAILED_NO_C_USER", "message": "No c_user cookie found. Account creation likely failed."}

    uid = session.cookies.get("c_user")
    profile_id = f'https://www.facebook.com/profile.php?id={uid}'

    cookie_dir = "/storage/emulated/0/cookie" # This path might need adjustment in a web server environment
    os.makedirs(cookie_dir, exist_ok=True)
    cookie_file = os.path.join(cookie_dir, f"{uid}.json")
    cookie_names = ["c_user", "datr", "fr", "noscript", "sb", "xs"]
    cookies_data = {name: session.cookies.get(name, "") for name in cookie_names}
    try:
        with open(cookie_file, "w") as f:
            json.dump(cookies_data, f, indent=4)
    except IOError as e:
        return {"status": "ERROR_SAVING_COOKIES", "message": f"Error saving cookies: {e}"}

    soup = BeautifulSoup(response.text, "html.parser")
    form_checkpoint = soup.find('form', action=lambda x: x and 'checkpoint' in x)
    if form_checkpoint:
        return {"status": "BLOCKED", "message": "Account created but blocked by checkpoint."}

    jbkj = None
    retries = 0
    while retries < MAX_RETRIES * 5:
        try:
            dtryvghjuijhn = requests.get(drtyghbj5hgcbv, timeout=30)
            dtryvghjuijhn.raise_for_status()
            soup_mail = BeautifulSoup(dtryvghjuijhn.text, "html.parser")
            table = soup_mail.find("table", class_="table table-hover table-striped")
            if table:
                subject_link = table.find("tbody", id="mail_list_body").find("a")
                if subject_link:
                    subject_div = subject_link.find("div")
                    if subject_div:
                        subject = subject_div.get_text(strip=True)
                        if "is your confirmation code" in subject:
                            jbkj = subject.replace(" is your confirmation code", "")
                            if jbkj:
                                break
        except requests.exceptions.RequestException as e:
            pass
        except Exception as e:
            pass
        time.sleep(5)
        retries += 1

    full_name = f"{firstname} {lastname}"
    access_token = ""

    filename_xlsx = "/storage/emulated/0/Acc_Created.xlsx" # This path might need adjustment
    filename_txt = "/storage/emulated/0/Acc_created.txt" # This path might need adjustment

    # In a Flask app, we'd typically acquire token immediately or through a separate user action
    # The interactive prompt "Do you want to save this account?" is removed.
    # The airplane mode prompt is also removed as it's not feasible in a web environment.

    # Try to get access token automatically
    if not has_access_token_in_xlsx(filename_xlsx, email_address):
        for _ in range(3):
            try:
                api_key = "882a8490361da98702bf97a021ddc14d"
                secret = "62f8ce9f74b12f84c123cc23437a4a32"
                params = {
                    "api_key": api_key,
                    "email": uid,
                    "format": "JSON",
                    "generate_session_cookies": 1,
                    "locale": "en_US",
                    "method": "auth.login",
                    "password": used_password,
                    "return_ssl_resources": 1,
                    "v": "1.0"
                }

                sig_str = "".join(f"{key}={params[key]}" for key in sorted(params)) + secret
                params["sig"] = hashlib.md5(sig_str.encode()).hexdigest()

                try:
                    resp = requests.get("https://api.facebook.com/restserver.php", params=params, headers=headers, timeout=60)
                    data = resp.json()
                    access_token = data.get("access_token", "")
                    if "error_title" in data:
                        return {"status": "TOKEN_ERROR", "message": f"Error acquiring token: {data['error_title']}"}
                except Exception as error_title:
                    return {"status": "TOKEN_EXCEPTION", "message": f"Exception acquiring token: {error_title}"}

                if access_token.strip():
                    data_to_save = [full_name, email_address, used_password, profile_id, access_token]
                    save_to_xlsx(filename_xlsx, data_to_save)
                    save_to_txt(filename_txt, data_to_save)
                    return {
                        "status": "SUCCESS",
                        "full_name": full_name,
                        "email": email_address,
                        "password": used_password,
                        "profile_id": uid,
                        "confirmation_code": jbkj if jbkj else 'N/A (Code not found)',
                        "access_token": access_token
                    }
                break
            except:
                pass
        else:
            return {"status": "NO_ACCESS_TOKEN", "message": "Failed to acquire access token."}
    else:
        return {"status": "ALREADY_HAS_TOKEN", "message": f"Account for {email_address} already has an access token."}


# --- Flask Routes ---

@app.route('/')
def index():
    return render_template_string("""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Facebook Account Creator</title>
            <style>
                :root {
                    /* Light Mode Defaults (Professional) */
                    --bg-color: #f8f9fa; /* Light grey/off-white */
                    --text-color: #343a40; /* Dark grey for readability */
                    --container-bg: #ffffff; /* Pure white */
                    --border-color: #e0e0e0; /* Light grey border */
                    --shadow-color: rgba(0, 0, 0, 0.08); /* Subtle shadow */
                    --button-bg: #007bff; /* Primary blue */
                    --button-hover-bg: #0056b3; /* Darker blue on hover */
                    --copy-button-bg: #28a745; /* Success green */
                    --copy-button-hover-bg: #218838; /* Darker green */
                    --results-bg: #e9ecef; /* Very light grey */
                    --results-border: #dee2e6; /* Light grey border */
                    --result-item-border: #cccccc; /* Slightly darker grey dash */
                    --header-color: #007bff; /* Primary blue for headers */
                    --status-success: #28a745; /* Green */
                    --status-failure: #dc3545; /* Red */
                    --status-warning: #ffc107; /* Yellow/Orange */
                    --input-focus-border: #80bdff; /* Blue for focus */
                    --input-focus-shadow: rgba(0, 123, 255, 0.25);
                }

                body.dark-mode {
                    /* Dark Mode (Professional) */
                    --bg-color: #2c3e50; /* Dark blue-grey */
                    --text-color: #ecf0f1; /* Light grey for text */
                    --container-bg: #34495e; /* Slightly darker container */
                    --border-color: #4a6572; /* Muted border */
                    --shadow-color: rgba(0, 0, 0, 0.3); /* More prominent shadow */
                    --button-bg: #1abc9c; /* Muted teal/green */
                    --button-hover-bg: #16a085; /* Darker teal */
                    --copy-button-bg: #3498db; /* Muted blue for copy */
                    --copy-button-hover-bg: #2980b9; /* Darker blue */
                    --results-bg: #4a6572; /* Darker grey-blue */
                    --results-border: #607d8b; /* Lighter border */
                    --result-item-border: #78909c; /* Lighter grey dash */
                    --header-color: #3498db; /* Muted blue for headers */
                    --status-success: #2ecc71; /* Brighter green */
                    --status-failure: #e74c3c; /* Brighter red */
                    --status-warning: #f39c12; /* Brighter orange */
                    --input-focus-border: #3498db; /* Muted blue for focus */
                    --input-focus-shadow: rgba(52, 152, 219, 0.35);
                }

                body {
                    font-family: 'Segoe UI', 'Roboto', 'Helvetica Neue', Arial, sans-serif; /* Professional font stack */
                    margin: 0;
                    padding: 20px;
                    background-color: var(--bg-color);
                    color: var(--text-color);
                    transition: background-color 0.3s ease, color 0.3s ease;
                    display: flex;
                    justify-content: center;
                    align-items: flex-start;
                    min-height: 100vh;
                }
                .container {
                    width: 100%;
                    max-width: 650px; /* Slightly wider for better content flow */
                    margin: 20px auto;
                    background-color: var(--container-bg);
                    padding: 30px;
                    border-radius: 8px;
                    box-shadow: 0 4px 12px var(--shadow-color); /* More pronounced shadow */
                    transition: background-color 0.3s ease, box-shadow 0.3s ease;
                }
                h1, h2 {
                    text-align: center;
                    color: var(--header-color);
                    margin-bottom: 25px;
                    font-weight: 600; /* Slightly bolder for headers */
                }
                label {
                    display: block;
                    margin-bottom: 8px;
                    font-weight: 500; /* Medium weight */
                    color: var(--text-color);
                }
                input[type="number"], input[type="text"], select {
                    width: calc(100% - 24px); /* Account for padding + border */
                    padding: 12px;
                    margin-bottom: 20px;
                    border: 1px solid var(--border-color);
                    border-radius: 5px; /* Slightly more rounded */
                    font-size: 16px;
                    background-color: var(--container-bg);
                    color: var(--text-color);
                    transition: border-color 0.3s ease, background-color 0.3s ease, color 0.3s ease, box-shadow 0.2s ease;
                    box-sizing: border-box; /* Include padding and border in element's total width and height */
                }
                input[type="number"]:focus, input[type="text"]:focus, select:focus {
                    outline: none;
                    border-color: var(--input-focus-border);
                    box-shadow: 0 0 0 0.2rem var(--input-focus-shadow);
                }
                button {
                    background-color: var(--button-bg);
                    color: white;
                    padding: 12px 25px; /* More padding for a better feel */
                    border: none;
                    border-radius: 5px;
                    cursor: pointer;
                    font-size: 17px;
                    font-weight: 600; /* Bolder button text */
                    width: 100%;
                    transition: background-color 0.3s ease, transform 0.2s ease;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }
                button:hover {
                    background-color: var(--button-hover-bg);
                    transform: translateY(-1px); /* Subtle lift effect */
                    box-shadow: 0 4px 8px rgba(0,0,0,0.15);
                }
                button:disabled {
                    background-color: #cccccc;
                    cursor: not-allowed;
                    transform: none;
                    box-shadow: none;
                }
                .copy-button {
                    background-color: var(--copy-button-bg);
                    color: white;
                    padding: 6px 12px; /* Adjusted padding */
                    border-radius: 4px; /* Slightly more rounded */
                    font-size: 13px;
                    margin-left: 15px; /* More space */
                    width: auto;
                    transition: background-color 0.3s ease, transform 0.2s ease;
                    box-shadow: 0 1px 2px rgba(0,0,0,0.1);
                }
                .copy-button:hover {
                    background-color: var(--copy-button-hover-bg);
                    transform: translateY(-1px);
                    box-shadow: 0 2px 4px rgba(0,0,0,0.15);
                }

                #results {
                    margin-top: 30px;
                    padding: 25px; /* More padding */
                    background-color: var(--results-bg);
                    border-radius: 8px;
                    border: 1px solid var(--results-border);
                    transition: background-color 0.3s ease, border-color 0.3s ease;
                }
                .result-item {
                    margin-bottom: 20px; /* More space between items */
                    padding-bottom: 20px;
                    border-bottom: 1px dashed var(--result-item-border);
                    transition: border-bottom-color 0.3s ease;
                }
                .result-item:last-child {
                    border-bottom: none;
                    margin-bottom: 0;
                    padding-bottom: 0;
                }
                .result-item p {
                    margin-bottom: 8px; /* Space between lines in a result item */
                }
                .status-success { color: var(--status-success); font-weight: bold; }
                .status-failure { color: var(--status-failure); font-weight: bold; }
                .status-warning { color: var(--status-warning); font-weight: bold; }

                /* Theme Toggle Switch Styles */
                .theme-switch-wrapper {
                    display: flex;
                    align-items: center;
                    justify-content: flex-end;
                    margin-bottom: 25px; /* More space below toggle */
                    font-size: 14px;
                }
                .theme-switch {
                    display: inline-block;
                    height: 24px; /* Slightly smaller toggle */
                    position: relative;
                    width: 48px; /* Adjusted width */
                    margin-right: 10px;
                }
                .theme-switch input {
                    display: none;
                }
                .slider {
                    background-color: #ccc;
                    bottom: 0;
                    cursor: pointer;
                    left: 0;
                    position: absolute;
                    right: 0;
                    top: 0;
                    transition: .4s;
                    border-radius: 24px; /* Fully rounded */
                }
                .slider:before {
                    background-color: #fff;
                    bottom: 2px; /* Adjusted for smaller size */
                    content: "";
                    height: 20px; /* Adjusted for smaller size */
                    left: 2px; /* Adjusted for smaller size */
                    position: absolute;
                    transition: .4s;
                    width: 20px; /* Adjusted for smaller size */
                    border-radius: 50%;
                }
                input:checked + .slider {
                    background-color: var(--button-bg); /* Use primary button color */
                }
                input:checked + .slider:before {
                    transform: translateX(24px); /* Adjusted translation */
                }
            </style>
        </head>
        <body>
            <div class="container">
                <div class="theme-switch-wrapper">
                    <label class="theme-switch" for="checkbox">
                        <input type="checkbox" id="checkbox" />
                        <span class="slider"></span>
                </div>

                <h1>Facebook Account Creator By Dars</h1>
                <form id="accountForm">
                    <label for="num_accounts">Number of Accounts to Create:</label>
                    <input type="number" id="num_accounts" name="num_accounts" value="5" min="1" required>

                    <label for="custom_password_base">Custom Password Base (Optional):</label>
                    <input type="text" id="custom_password_base" name="custom_password_base" placeholder="e.g., MyPass@">

                    <button type="submit" id="createButton">Create Accounts</button>
                </form>

                <div id="results">
                    <h2>Results:</h2>
                    <div id="resultsContent">
                        <p>No accounts created yet.</p>
                    </div>
                </div>
            </div>

            <script>
                // Theme Toggle Logic
                const toggleSwitch = document.getElementById('checkbox');
                const currentTheme = localStorage.getItem('theme');

                if (currentTheme) {
                    document.body.classList.add(currentTheme);
                    if (currentTheme === 'dark-mode') {
                        toggleSwitch.checked = true;
                    }
                } else if (window.matchMedia && window.matchMedia('(prefers-color-scheme: dark)').matches) {
                    // Check system preference if no local storage setting
                    document.body.classList.add('dark-mode');
                    toggleSwitch.checked = true;
                }

                function switchTheme(e) {
                    if (e.target.checked) {
                        document.body.classList.add('dark-mode');
                        localStorage.setItem('theme', 'dark-mode');
                    } else {
                        document.body.classList.remove('dark-mode');
                        localStorage.setItem('theme', 'light-mode');
                    }
                }
                toggleSwitch.addEventListener('change', switchTheme);


                // Function to copy text to clipboard
                function copyToClipboard(text, buttonElement) {
                    navigator.clipboard.writeText(text).then(function() {
                        const originalText = buttonElement.textContent;
                        buttonElement.textContent = 'Copied!';
                        setTimeout(() => {
                            buttonElement.textContent = originalText;
                        }, 1500);
                    }).catch(function(err) {
                        console.error('Could not copy text: ', err);
                        alert('Failed to copy. Please copy manually: ' + text);
                    });
                }

                document.getElementById('accountForm').addEventListener('submit', async function(event) {
                    event.preventDefault();

                    const createButton = document.getElementById('createButton');
                    const originalButtonText = createButton.textContent;
                    const resultsContent = document.getElementById('resultsContent');

                    createButton.textContent = 'Creating accounts... Please wait.';
                    createButton.disabled = true;

                    resultsContent.innerHTML = '<p>Initiating account creation process...</p>';

                    try {
                        const numAccounts = document.getElementById('num_accounts').value;
                        const customPasswordBase = document.getElementById('custom_password_base').value;

                        const response = await fetch('/create_accounts', {
                            method: 'POST',
                            headers: {
                                'Content-Type': 'application/json',
                            },
                            body: JSON.stringify({
                                num_accounts: parseInt(numAccounts),
                                custom_password_base: customPasswordBase
                            })
                        });

                        const data = await response.json();
                        resultsContent.innerHTML = '';

                        if (data.status === 'success') {
                            if (data.results && data.results.length > 0) {
                                data.results.forEach(account => {
                                    const accountDiv = document.createElement('div');
                                    accountDiv.classList.add('result-item');
                                    let statusClass = '';
                                    if (account.status === 'SUCCESS') {
                                        statusClass = 'status-success';
                                    } else if (account.status.startsWith('FAILED')) {
                                        statusClass = 'status-failure';
                                    } else if (account.status.includes('TOKEN') || account.status === 'BLOCKED') {
                                        statusClass = 'status-warning';
                                    }

                                    accountDiv.innerHTML = `
                                        <p><strong>Status:</strong> <span class="${statusClass}">${account.status}</span></p>
                                        <p><strong>Name:</strong> ${account.full_name || 'N/A'}</p>
                                        <p><strong>Email:</strong> <span id="email-${account.profile_id}">${account.email || 'N/A'}</span>
                                            <button class="copy-button" onclick="copyToClipboard(document.getElementById('email-${account.profile_id}').textContent, this)">Copy Email</button>
                                        </p>
                                        <p><strong>Password:</strong> <span id="password-${account.profile_id}">${account.password || 'N/A'}</span>
                                            <button class="copy-button" onclick="copyToClipboard(document.getElementById('password-${account.profile_id}').textContent, this)">Copy Password</button>
                                        </p>
                                        <p><strong>Confirmation Code:</strong> <span id="code-${account.profile_id}">${account.confirmation_code || 'N/A'}</span>
                                            <button class="copy-button" onclick="copyToClipboard(document.getElementById('code-${account.profile_id}').textContent, this)">Copy Code</button>
                                        </p>
                                        ${account.message ? `<p><strong>Message:</strong> ${account.message}</p>` : ''}
                                    `;
                                    resultsContent.appendChild(accountDiv);
                                });
                                // Auto-scroll to results
                                resultsContent.scrollIntoView({ behavior: 'smooth', block: 'start' });
                            } else {
                                resultsContent.innerHTML = '<p>No accounts were processed.</p>';
                            }
                        } else {
                            resultsContent.innerHTML = `<p class="status-failure">Error: ${data.message || 'Unknown error'}</p>`;
                        }

                    } catch (error) {
                        console.error('Error:', error);
                        resultsContent.innerHTML = `<p class="status-failure">An error occurred: ${error.message}</p>`;
                    } finally {
                        createButton.textContent = originalButtonText;
                        createButton.disabled = false;
                    }
                });
            </script>
        </body>
        </html>
    """)

@app.route('/create_accounts', methods=['POST'])
def create_accounts():
    data = request.get_json()
    num_accounts = data.get('num_accounts', 1)
    gender = random.choice([1, 2]) # 1 for male, 2 for female, now always random
    password_base = data.get('custom_password_base', None)

    if not isinstance(num_accounts, int) or num_accounts <= 0:
        return jsonify({"status": "error", "message": "Invalid number of accounts. Must be a positive integer."}), 400

    global custom_password_base
    custom_password_base = password_base

    results = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=num_accounts) as executor:
        futures = [
            executor.submit(create_fbunconfirmed_logic, i, "personal", gender)
            for i in range(1, num_accounts + 1)
        ]
        for future in concurrent.futures.as_completed(futures):
            results.append(future.result())

    return jsonify({"status": "success", "results": results})

if __name__ == '__main__':
    # Ensure necessary directories exist
    os.makedirs("/storage/emulated/0/cookie", exist_ok=True)
    app.run(debug=True, host='0.0.0.0', port=5000)
